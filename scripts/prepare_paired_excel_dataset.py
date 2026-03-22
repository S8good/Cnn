import argparse
import csv
import json
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
from openpyxl import load_workbook


PAIR_RE = re.compile(
    r"^\s*(?P<conc>[0-9]+(?:\.[0-9]+)?)\s*ng/ml\s*-\s*(?P<phase>bsa|ag)\s*-\s*(?P<rep>.+?)\s*$",
    re.IGNORECASE,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Prepare paired BSA/Ag Excel spectra into Stage-B few-shot dataset"
    )
    parser.add_argument("--input-xlsx", type=str, required=True, help="Paired spectra Excel path")
    parser.add_argument("--sheet", type=str, default="", help="Sheet name, default first sheet")
    parser.add_argument("--out-dir", type=str, default="data/real_fewshot_paired")
    parser.add_argument(
        "--representation",
        type=str,
        default="delta",
        choices=["delta", "ag", "bsa", "concat"],
        help="delta=Ag-BSA, ag=Ag only, bsa=BSA only, concat=[BSA,Ag]",
    )
    parser.add_argument(
        "--label-mode",
        type=str,
        default="concentration",
        choices=["concentration", "phase"],
        help="concentration: class by concentration; phase: class by representation/phase",
    )
    parser.add_argument("--target-start", type=float, default=400.0)
    parser.add_argument("--target-stop", type=float, default=800.0)
    parser.add_argument("--target-points", type=int, default=400)
    parser.add_argument("--no-resample", action="store_true")
    parser.add_argument("--normalize", type=str, default="minmax", choices=["none", "minmax", "zscore"])
    parser.add_argument("--drop-incomplete", action="store_true", help="Drop unmatched BSA/Ag pairs")
    return parser.parse_args()


def normalize_spectrum(x: np.ndarray, mode: str) -> np.ndarray:
    if mode == "none":
        return x
    if mode == "minmax":
        x_min = float(np.min(x))
        x_max = float(np.max(x))
        span = x_max - x_min
        if span <= 1e-12:
            return np.zeros_like(x)
        return (x - x_min) / span
    mean = float(np.mean(x))
    std = float(np.std(x))
    if std <= 1e-12:
        return np.zeros_like(x)
    return (x - mean) / std


def to_label_ids(labels: List[str]) -> Tuple[np.ndarray, Dict[str, int]]:
    uniq = sorted(set(labels))
    label_to_id = {name: i for i, name in enumerate(uniq)}
    ids = np.asarray([label_to_id[x] for x in labels], dtype=np.int64)
    return ids, label_to_id


def parse_column_name(name: str) -> Optional[Tuple[str, str, str]]:
    m = PAIR_RE.match(str(name))
    if not m:
        return None
    conc = m.group("conc")
    phase = m.group("phase").lower()
    rep = m.group("rep").strip()
    return conc, phase, rep


def build_vector(
    bsa: np.ndarray,
    ag: np.ndarray,
    representation: str,
) -> np.ndarray:
    if representation == "delta":
        return ag - bsa
    if representation == "ag":
        return ag
    if representation == "bsa":
        return bsa
    return np.concatenate([bsa, ag], axis=0)


def main() -> None:
    args = parse_args()
    in_xlsx = Path(args.input_xlsx)
    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if not in_xlsx.exists():
        raise FileNotFoundError(f"Excel not found: {in_xlsx}")

    wb = load_workbook(in_xlsx, read_only=True, data_only=True)
    sheet_name = args.sheet.strip() if args.sheet.strip() else wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
    header = next(rows)
    if header is None or len(header) < 3:
        raise ValueError("Invalid header, expected first col wavelength and many spectra columns")

    first_col = str(header[0]).strip().lower()
    if "wave" not in first_col:
        raise ValueError(f"First column should be wavelength, got: {header[0]}")

    col_meta: Dict[int, Tuple[str, str, str]] = {}
    for col_idx, name in enumerate(header[1:], start=1):
        parsed = parse_column_name(str(name))
        if parsed is not None:
            col_meta[col_idx] = parsed

    if not col_meta:
        raise ValueError("No paired spectrum columns detected. Please check column naming format.")

    wl_vals = []
    col_values: Dict[int, List[float]] = {idx: [] for idx in col_meta.keys()}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or row[0] is None:
            continue
        wl_vals.append(float(row[0]))
        for col_idx in col_meta.keys():
            v = row[col_idx] if col_idx < len(row) else None
            col_values[col_idx].append(float(v) if v is not None else np.nan)

    source_wl = np.asarray(wl_vals, dtype=np.float64)
    if len(source_wl) == 0:
        raise ValueError("No wavelength data found")

    if args.no_resample:
        target_wl = source_wl.copy()
    else:
        target_wl = np.linspace(args.target_start, args.target_stop, args.target_points, dtype=np.float64)

    pair_map: Dict[Tuple[str, str], Dict[str, np.ndarray]] = {}
    dropped_cols = 0
    for col_idx, (conc, phase, rep) in col_meta.items():
        raw = np.asarray(col_values[col_idx], dtype=np.float64)
        if np.any(np.isnan(raw)):
            dropped_cols += 1
            continue
        if args.no_resample:
            spec = raw
        else:
            spec = np.interp(target_wl, source_wl, raw)

        key = (conc, rep)
        pair_map.setdefault(key, {})[phase] = spec

    spectra: List[np.ndarray] = []
    labels: List[str] = []
    sample_ids: List[str] = []
    concentrations: List[float] = []
    phases: List[str] = []
    reps: List[str] = []
    dropped_pairs = 0

    for (conc, rep), entry in sorted(pair_map.items(), key=lambda x: (float(x[0][0]), x[0][1])):
        has_bsa = "bsa" in entry
        has_ag = "ag" in entry
        if not (has_bsa and has_ag):
            if args.drop_incomplete:
                dropped_pairs += 1
                continue
            raise ValueError(f"Incomplete pair for concentration={conc}, rep={rep}")

        bsa = entry["bsa"]
        ag = entry["ag"]
        vec = build_vector(bsa=bsa, ag=ag, representation=args.representation)
        vec = normalize_spectrum(vec, args.normalize).astype(np.float32)

        if args.label_mode == "concentration":
            label = f"{conc}ng/ml"
        else:
            label = "ag" if args.representation in ("ag", "delta", "concat") else "bsa"

        spectra.append(vec)
        labels.append(label)
        sample_ids.append(f"{conc}_{rep}")
        concentrations.append(float(conc))
        phases.append("paired")
        reps.append(rep)

    if not spectra:
        raise ValueError("No usable samples after pairing")

    spectra_arr = np.asarray(spectra, dtype=np.float32)
    label_ids, label_to_id = to_label_ids(labels)
    id_to_label = {int(v): k for k, v in label_to_id.items()}

    np.save(out_dir / "spectra.npy", spectra_arr)
    np.save(out_dir / "labels.npy", label_ids)
    np.save(out_dir / "wavelengths.npy", target_wl.astype(np.float32))

    with (out_dir / "metadata.csv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["sample_id", "label", "label_id", "concentration", "phase", "rep"])
        for i in range(len(sample_ids)):
            writer.writerow(
                [
                    sample_ids[i],
                    labels[i],
                    int(label_ids[i]),
                    concentrations[i],
                    phases[i],
                    reps[i],
                ]
            )

    payload = {
        "input_xlsx": str(in_xlsx.resolve()),
        "sheet": sheet_name,
        "representation": args.representation,
        "label_mode": args.label_mode,
        "normalize": args.normalize,
        "num_samples": int(len(labels)),
        "num_points": int(spectra_arr.shape[1]),
        "num_classes": int(len(label_to_id)),
        "label_to_id": label_to_id,
        "id_to_label": id_to_label,
        "source_wavelength_min": float(np.min(source_wl)),
        "source_wavelength_max": float(np.max(source_wl)),
        "target_wavelength_min": float(np.min(target_wl)),
        "target_wavelength_max": float(np.max(target_wl)),
        "target_points": int(len(target_wl)),
        "dropped_columns_with_nan": int(dropped_cols),
        "dropped_incomplete_pairs": int(dropped_pairs),
    }
    with (out_dir / "label_map.json").open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)

    print("Prepared paired real dataset:")
    print(f"  out_dir: {out_dir}")
    print(f"  samples: {len(labels)}")
    print(f"  classes: {len(label_to_id)}")
    print(f"  points : {spectra_arr.shape[1]}")


if __name__ == "__main__":
    main()
